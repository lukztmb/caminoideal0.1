from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Crear libro y hoja
wb = Workbook()
ws = wb.active
ws.title = "Menú Semanal"

# Encabezados de la tabla
headers = ["Día", "Desayuno", "Almuerzo", "Merienda", "Cena"]
ws.append(headers)

# Aplicar estilos al encabezado
header_font = Font(bold=True)
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)

    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Datos del menú semanal
menu = {
    "Lunes": [
        "Pan integral, huevo duro, mate o té", 
        "Milanesa de pollo al horno, ensalada (zanahoria, tomate, lechuga) y arroz pequeño",
        "Yogur natural y fruta", 
        "Omelette de 2 huevos con espinaca y cebolla, rodaja de pan opcional"
    ],
    "Martes": [
        "2 tostadas integrales con queso untable y café o té", 
        "Guiso de lentejas con verdura y carne picada (porción controlada); ensalada de repollo",
        "Banana y mate o infusión", 
        "Sopa casera de verduras, huevo duro y fruta"
    ],
    "Miércoles": [
        "Avena cocida con leche descremada o agua y 1/2 manzana rallada", 
        "Pollo al horno con calabaza y zanahoria; ensalada verde",
        "Fruta y un puñado de maní o nueces", 
        "Revuelto de huevo con acelga y cebolla; tostada"
    ],
    "Jueves": [
        "Yogur natural con avena y mate o té", 
        "Pastas con salsa casera de tomate y zanahoria rallada; ensalada de lechuga y tomate",
        "Fruta y té", 
        "Tortilla de papa (al horno, poca papa) y ensalada de rúcula"
    ],
    "Viernes": [
        "2 galletas de arroz con queso untable y café con leche descremada", 
        "Hamburguesa casera (lentejas o carne) con ensalada y 1 rebanada de pan",
        "Yogur y fruta", 
        "Pescado al horno con puré mixto (zapallo y papa)"
    ],
    "Sábado": [
        "Avena con banana pisada y té", 
        "Tarta de verdura casera (masa casera o comprada) y ensalada",
        "Infusión y tostada con mermelada light", 
        "Arroz con vegetales salteados y huevo"
    ],
    "Domingo": [

        "Pan con queso fresco e infusión", 
        "Asado o pollo al horno (porción magra), ensalada variada y papas al horno",
        "Gelatina light y fruta", 
        "Ensalada grande con atún, tomate, huevo y lechuga"
    ]
}

# Insertar los datos en la hoja
for dia, comidas in menu.items():
    row = [dia] + comidas
    ws.append(row)

# Ajustar el ancho de las columnas
ancho_columnas = {
    1: 12,  # Día
    2: 40,  # Desayuno
    3: 60,  # Almuerzo
    4: 35,  # Merienda
    5: 60   # Cena

}
for col, ancho in ancho_columnas.items():
    ws.column_dimensions[get_column_letter(col)].width = ancho

# Guardar el archivo
archivo = "menu_semanal.xlsx"
wb.save(archivo)
print(f"Archivo '{archivo}' generado exitosamente.")